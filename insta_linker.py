"""
╔══════════════════════════════════════════════════════════════════════════════╗
║          INSTAGRAM POST EXTRACTOR — PREMIUM TELEGRAM BOT v4                ║
║          + Views Count · Account Column · Private Log Channel              ║
║          + Cookie Health Check · Professional UI · Owner Credit            ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

#WORKING VERY WELL WITH VIEWS COUNT

import asyncio
import csv
import io
import json
import logging
import os
import random
import re
import time
import traceback
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum, auto
from io import BytesIO, StringIO
from pathlib import Path
from typing import Callable, Iterator, List, Optional, Tuple
from urllib.parse import urlparse

import requests
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from telegram import Update, ReplyKeyboardRemove
from telegram.constants import ParseMode
from telegram.ext import (
    Application, CommandHandler, ContextTypes,
    ConversationHandler, MessageHandler, filters,
)

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION (your values – unchanged)
# ══════════════════════════════════════════════════════════════════════════════

BOT_TOKEN           = "8689278721:AAHj7S0U0o65obtiRNmEb03W7sebGVDSllE"
IG_USERNAME         = "golangoverlord"
COOKIE_FILE         = r"C:\Users\asus\Downloads\www_instagram_com_cookies.txt"
PRIVATE_LOG_CHANNEL = "-1003867791896"
SESSION_DIR         = "./ig_sessions"
DELAY_MIN           = 1.5
DELAY_MAX           = 3.5
MAX_POSTS_SCAN      = 1000

# ══════════════════════════════════════════════════════════════════════════════
#  LOGGING — Console + Private Channel
# ══════════════════════════════════════════════════════════════════════════════

logging.basicConfig(
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S", level=logging.INFO,
)
log = logging.getLogger("ig_bot")
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("telegram").setLevel(logging.WARNING)

_app: Optional[Application] = None

async def log_to_channel(
    text: str = None,
    file=None,
    caption: str = None,
    parse_mode=ParseMode.HTML,
):
    """Send a message (and optionally a file) to the private log channel."""
    if not PRIVATE_LOG_CHANNEL or not _app:
        return
    try:
        if file:
            await _app.bot.send_document(
                chat_id=PRIVATE_LOG_CHANNEL,
                document=file,
                caption=caption[:1024] if caption else None,
                parse_mode=ParseMode.HTML,
            )
        else:
            await _app.bot.send_message(
                chat_id=PRIVATE_LOG_CHANNEL,
                text=text,
                parse_mode=ParseMode.HTML,
            )
    except Exception as e:
        log.error(f"Failed to send to log channel: {e}")

# ══════════════════════════════════════════════════════════════════════════════
#  DATA MODEL
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class PostRecord:
    date       : datetime   # UTC
    url        : str        # full permalink
    username   : str        # account owner
    view_count : Optional[int] = None   # only for videos/reels

# ══════════════════════════════════════════════════════════════════════════════
#  INSTAGRAM SESSION
# ══════════════════════════════════════════════════════════════════════════════

_SHORTCODE_ALPHABET = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-_"
_SHORTCODE_RE       = re.compile(r"instagram\.com/(?:p|reel|tv)/([A-Za-z0-9_\-]+)")

_session_cache: Optional[requests.Session] = None

def _parse_netscape_cookies(filepath: str) -> dict:
    cookies = {}
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split("\t")
            if len(parts) >= 7:
                name, value = parts[5].strip(), parts[6].strip()
                if name:
                    cookies[name] = value
    if "sessionid" not in cookies:
        raise RuntimeError("Cookie file missing 'sessionid'.")
    return cookies

def _build_session() -> requests.Session:
    global _session_cache
    if _session_cache is not None:
        return _session_cache

    log.info("Loading cookies from: %s", COOKIE_FILE)
    cookies = _parse_netscape_cookies(COOKIE_FILE)
    log.info("✓ %d cookies loaded", len(cookies))

    sess = requests.Session()
    sess.cookies.update(cookies)
    sess.headers.update({
        "User-Agent"      : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept"          : "*/*",
        "Accept-Language" : "en-US,en;q=0.9",
        "X-IG-App-ID"     : "936619743392459",
        "X-Requested-With": "XMLHttpRequest",
        "Referer"         : "https://www.instagram.com/",
        "Origin"          : "https://www.instagram.com",
    })

    csrf = cookies.get("csrftoken", "")
    if csrf:
        sess.headers["X-CSRFToken"] = csrf

    _session_cache = sess
    return sess

async def check_cookie_health() -> Tuple[bool, str]:
    """Test if the current session is still valid by fetching own profile."""
    try:
        sess = _build_session()
        url = "https://www.instagram.com/api/v1/accounts/current_user/?edit=true"
        resp = sess.get(url, timeout=15)
        if resp.status_code == 200:
            data = resp.json()
            username = data.get("user", {}).get("username", "unknown")
            return True, f"✅ Session is **valid** (logged in as @{username})"
        elif resp.status_code in (401, 403):
            return False, "❌ Session **expired** – please re‑export your cookies."
        else:
            return False, f"⚠️ Unexpected response (HTTP {resp.status_code})"
    except Exception as e:
        return False, f"❌ Error checking session: {e}"

# ══════════════════════════════════════════════════════════════════════════════
#  INSTAGRAM API HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _sleep():
    time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

def _shortcode_to_media_id(shortcode: str) -> int:
    n = 0
    for char in shortcode:
        n = n * 64 + _SHORTCODE_ALPHABET.index(char)
    return n

def _extract_shortcode(url: str) -> Optional[str]:
    m = _SHORTCODE_RE.search(url)
    return m.group(1) if m else None

def _api_get(sess: requests.Session, url: str, params: dict = None, retries=3) -> dict:
    for attempt in range(retries):
        try:
            resp = sess.get(url, params=params, timeout=20)
            if resp.status_code == 401:
                raise RuntimeError("Instagram returned 401 – session expired.")
            if resp.status_code == 429:
                raise RuntimeError("Rate limited (429). Wait a few minutes.")
            if resp.status_code != 200:
                raise RuntimeError(f"HTTP {resp.status_code}")
            return resp.json()
        except (requests.RequestException, json.JSONDecodeError) as e:
            if attempt == retries - 1:
                raise RuntimeError(f"API request failed after {retries} attempts: {e}")
            wait = 2 ** attempt + random.random()
            time.sleep(wait)
    raise RuntimeError("Unexpected error in _api_get")

def _get_post_info(sess: requests.Session, shortcode: str) -> dict:
    media_id = _shortcode_to_media_id(shortcode)
    url = f"https://i.instagram.com/api/v1/media/{media_id}/info/"
    data = _api_get(sess, url)
    items = data.get("items") or []
    if not items:
        raise RuntimeError(f"No data for shortcode: {shortcode}")
    return items[0]

def _get_user_id(sess: requests.Session, username: str) -> str:
    url = "https://www.instagram.com/api/v1/users/web_profile_info/"
    data = _api_get(sess, url, params={"username": username})
    try:
        return str(data["data"]["user"]["id"])
    except (KeyError, TypeError):
        raise RuntimeError(f"Could not resolve user ID for @{username}")

def _iter_user_posts(
    sess: requests.Session,
    user_id: str,
    stop_before_ts: int,
    progress: Callable[[str], None] = None,
) -> Iterator[dict]:
    url = f"https://i.instagram.com/api/v1/feed/user/{user_id}/"
    max_id = None
    scanned = 0

    while True:
        params = {"count": 12}
        if max_id:
            params["max_id"] = max_id

        data = _api_get(sess, url, params=params)
        items = data.get("items") or []

        if not items:
            break

        for item in items:
            scanned += 1
            ts = item.get("taken_at", 0)

            if ts < stop_before_ts:
                return

            if scanned > MAX_POSTS_SCAN:
                if progress:
                    progress(f"⚠️ Hit {MAX_POSTS_SCAN} post limit – stopping.")
                return

            yield item
            _sleep()

        next_max = data.get("next_max_id")
        if not next_max or not data.get("more_available"):
            break
        max_id = next_max
        _sleep()

# ══════════════════════════════════════════════════════════════════════════════
#  SCRAPER — IMPROVED VIEW COUNT EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def scrape_posts_between(
    start_url: str,
    end_url: str,
    on_progress: Optional[Callable[[str], None]] = None,
) -> List[PostRecord]:
    def _prog(msg: str):
        log.info(msg)
        if on_progress:
            on_progress(msg)

    sc_start = _extract_shortcode(start_url)
    sc_end   = _extract_shortcode(end_url)
    if not sc_start: raise ValueError(f"Invalid start URL: {start_url}")
    if not sc_end:   raise ValueError(f"Invalid end URL: {end_url}")
    if sc_start == sc_end: raise ValueError("Start and End are the same post.")

    sess = _build_session()

    _prog("🔍 Loading anchor posts …")
    info_start = _get_post_info(sess, sc_start)
    _sleep()
    info_end   = _get_post_info(sess, sc_end)

    ts_start = info_start["taken_at"]
    ts_end   = info_end["taken_at"]
    user_start = info_start["user"]["username"]
    user_end   = info_end["user"]["username"]

    if user_start != user_end:
        raise ValueError(f"Posts belong to different profiles: @{user_start} vs @{user_end}")

    username = user_start

    if ts_start > ts_end:
        ts_start, ts_end = ts_end, ts_start
        sc_start, sc_end = sc_end, sc_start
        log.info("Anchor order reversed – swapped automatically.")

    dt_start = datetime.fromtimestamp(ts_start, tz=timezone.utc)
    dt_end   = datetime.fromtimestamp(ts_end,   tz=timezone.utc)
    _prog(f"📅 @{username}  |  {dt_start.date()} → {dt_end.date()}")

    _prog("📋 Resolving user ID …")
    user_id = _get_user_id(sess, username)
    log.info("User ID: %s", user_id)

    _prog("📋 Scanning posts …")
    collected: List[PostRecord] = []

    # Possible keys that might contain view count
    VIEW_KEYS = ['video_view_count', 'play_count', 'view_count']

    for item in _iter_user_posts(sess, user_id, stop_before_ts=ts_start, progress=_prog):
        ts = item["taken_at"]
        if ts > ts_end:
            continue

        shortcode = item.get("code", "")
        dt = datetime.fromtimestamp(ts, tz=timezone.utc)

        # --- Robust view count extraction ---
        view_count = None
        media_type = item.get("media_type")  # 1=photo, 2=video, 8=carousel

        # Try main item if it's a video (type 2) or carousel (type 8 – we'll check inside later)
        if media_type == 2:
            for key in VIEW_KEYS:
                if key in item and item[key] is not None:
                    view_count = item[key]
                    break
        elif media_type == 8:
            # Carousel: search for the first video inside
            for media in item.get("carousel_media", []):
                if media.get("media_type") == 2:
                    for key in VIEW_KEYS:
                        if key in media and media[key] is not None:
                            view_count = media[key]
                            break
                    if view_count is not None:
                        break
        # If media_type is 1 (photo) or not found, view_count stays None

        collected.append(PostRecord(
            date=dt,
            url=f"https://www.instagram.com/p/{shortcode}/",
            username=username,
            view_count=view_count,
        ))

        if len(collected) % 10 == 0:
            _prog(f"📦 {len(collected)} posts collected …")

    collected.reverse()
    _prog(f"✅ Done: {len(collected)} posts.")
    return collected

# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

_C = {
    "bg_dark": "0D0D0D", "bg_mid": "1A1A2E",
    "row_a"  : "16213E", "row_b" : "0F3460",
    "accent" : "E94560", "gold"  : "F5A623", "white": "FFFFFF",
}

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(c=_C["white"], bold=False, sz=10):
    return Font(bold=bold, color=c, size=sz, name="Consolas")

_SIDE   = Side(style="thin", color="2A2A4A")
_BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
_MID    = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center")

def _generate_excel(records: List[PostRecord]) -> BytesIO:
    wb = openpyxl.Workbook()

    # Posts sheet
    ws = wb.active
    ws.title = "Posts"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _C["accent"]

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "⚡  INSTAGRAM POST EXPORT  —  @androbeet"
    c.font = _font(_C["accent"], bold=True, sz=13)
    c.fill = _fill(_C["bg_dark"])
    c.alignment = _MID
    ws.row_dimensions[1].height = 28

    headers = ["  #", "  Date (DD/MM/YYYY)", "  Time (UTC)", "  Account", "  Link", "  Views"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.font = _font(_C["accent"], bold=True)
        c.fill = _fill(_C["bg_mid"])
        c.alignment = _MID
        c.border = _BORDER
    ws.row_dimensions[2].height = 22

    for i, rec in enumerate(records):
        ri = i + 3
        bg = _C["row_a"] if i % 2 == 0 else _C["row_b"]
        ws.cell(ri, 1, i + 1).alignment = _MID
        ws.cell(ri, 2, rec.date.strftime("%d/%m/%Y")).alignment = _MID
        ws.cell(ri, 3, rec.date.strftime("%H:%M")).alignment = _MID
        ws.cell(ri, 4, f"@{rec.username}").alignment = _LEFT
        ws.cell(ri, 5, rec.url).alignment = _LEFT
        ws.cell(ri, 6, rec.view_count if rec.view_count is not None else "").alignment = _MID

        for col in range(1, 7):
            c = ws.cell(ri, col)
            c.font = _font()
            c.fill = _fill(bg)
            c.border = _BORDER
        ws.row_dimensions[ri].height = 17

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 58
    ws.column_dimensions["F"].width = 10
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:F{len(records) + 2}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = _C["gold"]

    total_views = sum(r.view_count for r in records if r.view_count)
    rows = [
        ("EXPORT SUMMARY",          ""),
        ("Generated (UTC)",         datetime.utcnow().strftime("%d %b %Y  %H:%M")),
        ("Total Posts",             str(len(records))),
        ("Total Views (videos)",    str(total_views) if total_views else "N/A"),
    ]
    if records:
        rows += [
            ("Earliest Post",        records[0].date.strftime("%d/%m/%Y %H:%M")),
            ("Latest Post",          records[-1].date.strftime("%d/%m/%Y %H:%M")),
            ("Date Range (days)",    str((records[-1].date - records[0].date).days)),
        ]

    for ri, (label, value) in enumerate(rows, 2):
        lc = ws2.cell(ri, 2, label)
        vc = ws2.cell(ri, 3, value)
        if label == "EXPORT SUMMARY":
            lc.font = _font(_C["gold"], bold=True, sz=13)
            ws2.merge_cells(f"B{ri}:C{ri}")
            lc.alignment = _MID
        else:
            lc.font = _font(_C["accent"], bold=True)
            vc.font = _font(_C["white"])
            lc.alignment = Alignment(horizontal="right", vertical="center")
            vc.alignment = _LEFT
        for c in (lc, vc):
            c.fill = _fill(_C["bg_mid"])
            c.border = _BORDER
        ws2.row_dimensions[ri].height = 22

    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 28

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ══════════════════════════════════════════════════════════════════════════════
#  CSV GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

def _generate_csv(records: List[PostRecord]) -> BytesIO:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["#", "Date (DD/MM/YYYY)", "Time (UTC)", "Account", "Link", "Views"])
    for i, rec in enumerate(records, 1):
        writer.writerow([
            i,
            rec.date.strftime("%d/%m/%Y"),
            rec.date.strftime("%H:%M"),
            f"@{rec.username}",
            rec.url,
            rec.view_count if rec.view_count is not None else "",
        ])
    buf = BytesIO()
    buf.write(output.getvalue().encode("utf-8"))
    buf.seek(0)
    return buf

# ══════════════════════════════════════════════════════════════════════════════
#  TELEGRAM BOT — Conversation & Commands
# ══════════════════════════════════════════════════════════════════════════════

class S(Enum):
    WAIT_START = auto()
    WAIT_END   = auto()
    WAIT_FORMAT = auto()

_WELCOME = """
╔══════════════════════════════════╗
║   📸  INSTAGRAM POST EXTRACTOR  ║
╚══════════════════════════════════╝
                                        created by **@androbeet**
I'll collect every post / reel between two anchor posts and send you a clean file with **views count** and **account name**.

──────────────────────────────────
*Step 1 of 2* — Send the *Start URL*
_(the older / earlier post)_
──────────────────────────────────

Example:
`https://www.instagram.com/p/ABC123/`

Send /cancel to abort.
""".strip()

_STEP2 = """
✅ *Start URL saved!*

──────────────────────────────────
*Step 2 of 2* — Send the *End URL*
_(the newer / more recent post)_
──────────────────────────────────
""".strip()

_HELP = """
📖 *HOW TO USE* — created by **@androbeet**

1️⃣  /start  
2️⃣  Paste the *older* post URL  
3️⃣  Paste the *newer* post URL  
4️⃣  Receive your `.xlsx` file 🎉  

*Accepted formats:*  
• `instagram.com/p/SHORTCODE/`  
• `instagram.com/reel/SHORTCODE/`  

*New in this version:*  
• Views count & account column in Excel  
• Cookie health check via /check  
• All exports logged to private channel  

/cancel — abort any time
""".strip()

def _is_ig_url(text: str) -> bool:
    return bool(_SHORTCODE_RE.search(text))

async def h_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> S:
    ctx.user_data.clear()
    await update.message.reply_text(_WELCOME, parse_mode=ParseMode.HTML)
    return S.WAIT_START

async def h_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(_HELP, parse_mode=ParseMode.HTML)

async def h_cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    ctx.user_data.clear()
    await update.message.reply_text(
        "🚫 *Cancelled.* Send /start whenever you're ready.",
        parse_mode=ParseMode.HTML,
    )
    return ConversationHandler.END

async def h_check(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    msg = await update.message.reply_text("🔍 Checking cookie health...")
    valid, status = await check_cookie_health()
    await msg.edit_text(status, parse_mode=ParseMode.HTML)

async def h_receive_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> S:
    url = update.message.text.strip()
    if not _is_ig_url(url):
        await update.message.reply_text(
            "❌ Not a valid Instagram URL.\n\nExample:\n"
            "`https://www.instagram.com/p/ABC123/`",
            parse_mode=ParseMode.HTML,
        )
        return S.WAIT_START
    ctx.user_data["start_url"] = url
    await update.message.reply_text(_STEP2, parse_mode=ParseMode.HTML)
    return S.WAIT_END

async def h_receive_end(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    url = update.message.text.strip()
    start_url = ctx.user_data.get("start_url", "")

    if not _is_ig_url(url):
        await update.message.reply_text(
            "❌ Not a valid Instagram URL.\n\nExample:\n"
            "`https://www.instagram.com/reel/XYZ789/`",
            parse_mode=ParseMode.HTML,
        )
        return S.WAIT_END

    prog = await update.message.reply_text("⏳ *Starting …*", parse_mode=ParseMode.HTML)
    loop = asyncio.get_running_loop()

    async def upd(text: str):
        try:
            await prog.edit_text(text, parse_mode=ParseMode.HTML)
        except Exception:
            pass

    def progress_sync(msg: str):
        asyncio.run_coroutine_threadsafe(upd(msg), loop)

    try:
        await update.effective_chat.send_action("typing")
        records: List[PostRecord] = await loop.run_in_executor(
            None,
            lambda: scrape_posts_between(
                start_url, url,
                on_progress=progress_sync,
            ),
        )

        if not records:
            await upd("⚠️ *No posts found* between those two links.")
            return ConversationHandler.END

            await upd(f"✅ *{len(records)} posts found* — generating files …")

        # FILES LOG MEIN JAYENGI IG TXT FILE MEIN
        csv_buf = _generate_csv(records)

        
        excel_buf = None
        try:
            excel_buf = _generate_excel(records)
        except Exception as e:
            log.error(f"Excel generation failed: {e}")
            await update.message.reply_text("⚠️ Excel generation failed. Sending CSV instead.")

        caption = (
            "📊 <b>INSTAGRAM POST EXPORT</b>\n"
            "────────────────────────\n"
            f"📦 Posts : <b>{len(records)}</b>\n"
            f"👤 Account: @{records[0].username}\n"
            f"📅 From  : <code>{records[0].date.strftime('%d %b %Y %H:%M')}</code>\n"
            f"📅 To    : <code>{records[-1].date.strftime('%d %b %Y %H:%M')}</code>\n"
            f"🎬 Views : <b>{sum(r.view_count for r in records if r.view_count)}</b> (videos only)\n"
            "────────────────────────\n"
            "<i>/start for another export</i>"
        )

        
        if excel_buf:
            await update.message.reply_document(
                document=excel_buf,
                filename=f"instagram_export_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx",
                caption=caption,
                parse_mode=ParseMode.HTML,
            )
            excel_buf.seek(0)          
        else:
            await update.message.reply_document(
                document=csv_buf,
                filename=f"instagram_export_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.csv",
                caption=caption,
                parse_mode=ParseMode.HTML,
            )
            csv_buf.seek(0)             

       
        if excel_buf:
            await log_to_channel(
                text=None,
                file=excel_buf,
                caption=f"📤 Export by @{update.effective_user.username or update.effective_user.id}\n{caption}",
            )
        # CSV is always sent to the log channel
        await log_to_channel(
            text=None,
            file=csv_buf,
            caption=f"📄 CSV Export – @{records[0].username}",
        )

        await prog.delete()

    except RuntimeError as exc:
        log.error("Instagram API error: %s", exc)
        await upd(f"🔐 *Instagram Error:*\n\n{exc}")
        await log_to_channel(f"❌ Error for user {update.effective_user.id}: {exc}")

    except ValueError as exc:
        log.warning("Validation error: %s", exc)
        await upd(f"❌ *Error:*\n\n{exc}")

    except Exception:
        log.error("Unhandled:\n%s", traceback.format_exc())
        await upd("💥 *Unexpected error.* Check the terminal for details.")

    finally:
        ctx.user_data.clear()

    return ConversationHandler.END

# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main():
    log.info("━" * 47)
    log.info("  Instagram Post Extractor Bot v4")
    log.info("  Cookies : %s", COOKIE_FILE)
    log.info("  Log Channel : %s", PRIVATE_LOG_CHANNEL or "Disabled")
    log.info("━" * 47)

    global _app
    _app = Application.builder().token(BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", h_start)],
        states={
            S.WAIT_START: [MessageHandler(filters.TEXT & ~filters.COMMAND, h_receive_start)],
            S.WAIT_END:   [MessageHandler(filters.TEXT & ~filters.COMMAND, h_receive_end)],
        },
        fallbacks=[CommandHandler("cancel", h_cancel)],
        allow_reentry=True,
    )
    _app.add_handler(conv)
    _app.add_handler(CommandHandler("help", h_help))
    _app.add_handler(CommandHandler("check", h_check))

    log.info("Bot is polling. Ctrl+C to stop.")
    _app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
